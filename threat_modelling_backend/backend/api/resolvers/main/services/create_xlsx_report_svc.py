# Copyright 2024 Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: LicenseRef-.amazon.com.-AmznSL-1.0
# Licensed under the Amazon Software License  http://aws.amazon.com/asl/

import boto3
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import os
import pandas as pd
from typing import TYPE_CHECKING

from aws_lambda_powertools import Logger

from genai_core.model import ThreatModel

if TYPE_CHECKING:
    from mypy_boto3_s3.client import S3Client
else:
    S3Client = object

DATA_BUCKET_NAME = os.getenv("DATA_BUCKET_NAME")

logger = Logger()

s3_client: S3Client = boto3.session.Session(region_name=os.getenv("AWS_REGION", "us-east-1")).client("s3")


def create_xlsx_report(threat_model: ThreatModel) -> io.BytesIO:
    threat_model_df = pd.json_normalize(
        threat_model.model_dump(),
        record_path=['diagrams', 'components', 'threats'],  # Path to the deepest level (C objects)
        meta=[
            ['diagrams', 'components', 'name'],
            ['diagrams', 'components', 'description'],
            ['diagrams', 'components', 'component_type'],
        ],
        errors='ignore'
    )

    # let's create a new column to check if a certain threat has been evaluated. To us, being "evaluated" means there is a reason for the action
    threat_model_df["evaluated"] = threat_model_df["reason"].apply(
        lambda reason: "Evaluated" if reason else "Not Evaluated")

    # let's create a new columns with the sum of dread scores. We will use this to gauge risk of the threat model using a histogram
    threat_model_df["total_dread_score"] = (
            threat_model_df["dread_scores.damage"] + threat_model_df["dread_scores.reproducibility"] +
            threat_model_df["dread_scores.exploitability"] + threat_model_df["dread_scores.affected_users"] +
            threat_model_df["dread_scores.discoverability"]
    )

    score_bins = [1, 10, 20, 30, 40, 50]
    score_bin_labels = ["1-10", "10-20", "20-30", "30-40", "40-50"]
    threat_model_df["dread_score_bins"] = pd.cut(threat_model_df["total_dread_score"], bins=score_bins,
                                                 labels=score_bin_labels)
    score_bin_counts = threat_model_df["dread_score_bins"].value_counts(sort=False)

    summary_by_action = pd.pivot_table(threat_model_df[["action", "evaluated", "id"]],
                                       index='action',
                                       columns='evaluated',
                                       values='id', aggfunc='count')

    output = io.BytesIO()
    # output = "pivot_table.xlsx"

    # Create an Excel writer object
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_by_action.to_excel(writer, sheet_name='Summary by action')

        score_bin_counts_df = pd.DataFrame({"Bins": score_bin_counts.index, "Frenquency": score_bin_counts.values})
        score_bin_counts_df.to_excel(writer, sheet_name='Dread Score Frenquency', index=False)

        # All data
        threat_model_df.to_excel(writer, sheet_name='Full Data', index=False)

    # output.seek(0)
    wb = load_workbook(output)
    ws = wb["Summary by action"]

    ws[
        "E2"] = "Currently we assume that threats without a reason have not yet been evaluated. The goal is for every threat to have an action-reason filled."

    ws = wb["Dread Score Frenquency"]

    chart = BarChart()
    chart.title = "Histogram of DREAD SCORES sum"
    chart.x_axis.title = "Sum Range"
    chart.y_axis.title = "Frequency"

    # Reference the bin labels (Bins) and frequencies for the histogram
    bin_data = Reference(ws, min_col=2, min_row=2, max_row=len(score_bin_counts_df) + 1)  # Frequency column
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(score_bin_counts_df) + 1)  # Bins column

    # Add the data and categories to the chart
    chart.add_data(bin_data, titles_from_data=True)
    chart.set_categories(categories)

    # Place the chart on the worksheet
    ws.add_chart(chart, "E5")  # Position the chart at cell E5
    ws[
        "E2"] = "\"Riskier\"  Threat Models will lean heavy on towards the right in the histogram chart, indicating they have threats with higher DREAD scores."

    # Save the workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_report(threat_model: ThreatModel) -> str:
    generated_report = create_xlsx_report(threat_model)

    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M")

    file_name = f"reports/{current_time}-{threat_model.id}.xlsx"

    s3_client.upload_fileobj(generated_report, Bucket=DATA_BUCKET_NAME, Key=file_name)

    presigned_url = s3_client.generate_presigned_url("get_object",
                                                     Params={"Bucket": DATA_BUCKET_NAME, "Key": file_name},
                                                     ExpiresIn=3600)

    return presigned_url
